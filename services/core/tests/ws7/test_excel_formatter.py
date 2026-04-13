"""
Tests for WS7 Excel Formatter — WestMetro formatting compliance.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook
import io

from src.reports.excel_formatter import (
    HEADER_BG,
    SUBHEADER_BG,
    SNO_BG,
    ALT_ROW,
    WHITE,
    AlignmentCategory,
    ColumnConfig,
    add_sheet,
    apply_column_widths,
    apply_data_borders,
    apply_freeze_panes,
    apply_white_space,
    create_workbook,
    format_sheet_complete,
    save_workbook,
    write_data_rows,
    write_headers,
)


def _sample_columns():
    return [
        ColumnConfig("S/No", "sno", width=7, is_sno=True),
        ColumnConfig("Invoice ID", "invoice_id", width=21, alignment=AlignmentCategory.ID_SHORT),
        ColumnConfig("Amount \u20a6", "amount", width=23, alignment=AlignmentCategory.MONETARY, number_format="#,##0.00", is_monetary=True),
        ColumnConfig("Customer", "customer", width=40, alignment=AlignmentCategory.DESCRIPTION),
    ]


def _sample_rows():
    return [
        {"sno": 1, "invoice_id": "INV-001", "amount": 10000.50, "customer": "Acme Corp"},
        {"sno": 2, "invoice_id": "INV-002", "amount": None, "customer": "Beta Ltd"},
        {"sno": 3, "invoice_id": "INV-003", "amount": 25000.00, "customer": "Gamma Inc"},
    ]


class TestWorkbookLifecycle:
    def test_create_removes_default_sheet(self):
        wb = create_workbook()
        assert len(wb.sheetnames) == 0

    def test_add_sheet(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test Sheet")
        assert "Test Sheet" in wb.sheetnames
        assert ws.title == "Test Sheet"

    def test_save_returns_bytes(self):
        wb = create_workbook()
        add_sheet(wb, "Data")
        content = save_workbook(wb)
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_saved_workbook_is_valid_xlsx(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Data")
        ws.cell(row=1, column=1, value="Test")
        content = save_workbook(wb)
        # Re-load to verify
        loaded = load_workbook(io.BytesIO(content))
        assert "Data" in loaded.sheetnames


class TestHeaders:
    def test_single_row_header(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        start_row = write_headers(ws, cols, schema_paths=None)
        assert start_row == 2  # Data starts at row 2

        # Check header cells
        assert ws.cell(row=1, column=1).value == "S/No"
        assert ws.cell(row=1, column=2).value == "Invoice ID"
        assert ws.cell(row=1, column=3).value == "Amount \u20a6"
        assert ws.cell(row=1, column=4).value == "Customer"

    def test_two_row_header(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        paths = ["sno", "invoice_id", "amount", "customer_name"]
        start_row = write_headers(ws, cols, schema_paths=paths)
        assert start_row == 4  # Data starts at row 4

        # Row 1: display names
        assert ws.cell(row=1, column=1).value == "S/No"
        # Row 2: schema paths
        assert ws.cell(row=2, column=1).value == "sno"

    def test_header_fill_color(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        write_headers(ws, _sample_columns())
        cell = ws.cell(row=1, column=1)
        assert cell.fill.start_color.rgb == "00" + HEADER_BG


class TestDataRows:
    def test_write_rows(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        rows = _sample_rows()
        last_row = write_data_rows(ws, rows, cols, start_row=2)
        assert last_row == 4  # 3 rows starting at row 2

    def test_null_handling_dash(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        rows = [{"sno": 1, "invoice_id": "INV-001", "amount": None, "customer": None}]
        write_data_rows(ws, rows, cols, start_row=2)
        # amount and customer should be "-"
        assert ws.cell(row=2, column=3).value == "-"
        assert ws.cell(row=2, column=4).value == "-"

    def test_sno_always_sno_fill(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        rows = _sample_rows()
        write_data_rows(ws, rows, cols, start_row=2)
        # All S/No cells should have SNO_BG fill
        for row_num in range(2, 5):
            cell = ws.cell(row=row_num, column=1)
            assert cell.fill.start_color.rgb == "00" + SNO_BG

    def test_zebra_striping(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        rows = _sample_rows()
        write_data_rows(ws, rows, cols, start_row=2)
        # Row 2 (offset 0): not alt → WHITE
        # Row 3 (offset 1): alt → ALT_ROW
        # (Check non-S/No column)
        assert ws.cell(row=2, column=2).fill.start_color.rgb == "00" + WHITE
        assert ws.cell(row=3, column=2).fill.start_color.rgb == "00" + ALT_ROW

    def test_empty_rows(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        last = write_data_rows(ws, [], _sample_columns(), start_row=2)
        assert last == 1  # start_row - 1


class TestFormatSheetComplete:
    def test_full_pipeline(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        format_sheet_complete(ws, _sample_columns(), _sample_rows())
        # Verify workbook is valid
        content = save_workbook(wb)
        loaded = load_workbook(io.BytesIO(content))
        ws2 = loaded["Test"]
        # Header row exists
        assert ws2.cell(row=1, column=1).value == "S/No"
        # Data row 1 at row 2 (single-row header, no schema_paths)
        assert ws2.cell(row=2, column=1).value == 1

    def test_freeze_panes_set(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        format_sheet_complete(ws, _sample_columns(), _sample_rows())
        assert ws.freeze_panes is not None


class TestColumnWidths:
    def test_widths_applied(self):
        wb = create_workbook()
        ws = add_sheet(wb, "Test")
        cols = _sample_columns()
        rows = _sample_rows()
        write_headers(ws, cols)
        write_data_rows(ws, rows, cols, start_row=4)
        apply_column_widths(ws, cols, len(rows), start_row=4)
        # S/No should be at least 7
        from openpyxl.utils import get_column_letter
        assert ws.column_dimensions[get_column_letter(1)].width >= 7
