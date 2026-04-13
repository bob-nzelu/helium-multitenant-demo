"""
Tests for WS7 Report Generators — verify each generator produces valid output.

PDF generators: verify HTML content (not PDF rendering — that's weasyprint's job).
Excel generators: verify openpyxl workbook structure and data.
"""

from __future__ import annotations

import io
from datetime import date, datetime, timezone
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook

from src.reports.generators import compliance, transmission, customer, audit_trail, monthly_summary


# ── Helpers ──────────────────────────────────────────────────────────────


def _make_mock_pool(query_results: dict[str, list] | None = None):
    """Create a mock pool that returns predefined query results."""
    pool = AsyncMock()
    conn = AsyncMock()

    call_count = [0]
    results_queue = list((query_results or {}).values())

    cursor = AsyncMock()

    async def mock_execute(query, params=None):
        nonlocal call_count
        c = AsyncMock()
        # Return different results for different queries
        if results_queue and call_count[0] < len(results_queue):
            result = results_queue[call_count[0]]
            c.fetchone = AsyncMock(return_value=result[0] if result else None)
            c.fetchall = AsyncMock(return_value=result)
            c.description = [type("D", (), {"name": f"col{i}"})() for i in range(20)]
        else:
            c.fetchone = AsyncMock(return_value=None)
            c.fetchall = AsyncMock(return_value=[])
            c.description = []
        call_count[0] += 1
        return c

    conn.execute = mock_execute

    cm = AsyncMock()
    cm.__aenter__ = AsyncMock(return_value=conn)
    cm.__aexit__ = AsyncMock(return_value=False)
    pool.connection = MagicMock(return_value=cm)
    return pool


# ── Compliance Generator ─────────────────────────────────────────────────


class TestComplianceGenerator:
    @pytest.mark.asyncio
    async def test_returns_bytes_and_content_type(self):
        pool = _make_mock_pool({
            "metrics": [(100, 95, 5, 3, 2)],
            "reasons": [],
            "trend": [],
        })

        with patch("src.reports.generators.compliance._compliance", new_callable=AsyncMock) as mock_comp:
            mock_comp.return_value = {
                "firs_acceptance_rate": 95.0,
                "submitted_count": 100,
                "accepted_count": 95,
                "rejected_count": 5,
                "overdue_count": 3,
                "disputed_count": 2,
                "common_rejection_reasons": [
                    {"reason": "Invalid TIN", "count": 3},
                ],
                "monthly_trend": [
                    {"month": "2026-01", "score": 90.0},
                    {"month": "2026-02", "score": 95.0},
                ],
            }

            content, ct = await compliance.generate(
                pool,
                {"date_from": "2026-01-01", "date_to": "2026-03-31"},
                "comp-1",
            )

        assert isinstance(content, bytes)
        assert ct == "application/pdf"
        # Since weasyprint likely not installed, verify HTML content
        html = content.decode("utf-8", errors="replace")
        assert "Compliance Report" in html
        assert "95.0%" in html or "95%" in html

    @pytest.mark.asyncio
    async def test_rejection_table_rendered(self):
        with patch("src.reports.generators.compliance._compliance", new_callable=AsyncMock) as mock_comp:
            mock_comp.return_value = {
                "firs_acceptance_rate": 80.0,
                "submitted_count": 50,
                "accepted_count": 40,
                "rejected_count": 10,
                "overdue_count": 0,
                "disputed_count": 0,
                "common_rejection_reasons": [
                    {"reason": "Invalid TIN", "count": 7},
                    {"reason": "Missing HS code", "count": 3},
                ],
                "monthly_trend": [],
            }

            content, _ = await compliance.generate(
                _make_mock_pool(), {}, "comp-1",
            )

        html = content.decode("utf-8", errors="replace")
        assert "Invalid TIN" in html
        assert "Missing HS code" in html


# ── Transmission Generator ───────────────────────────────────────────────


class TestTransmissionGenerator:
    @pytest.mark.asyncio
    async def test_returns_valid_xlsx(self):
        # Mock DB rows
        pool = AsyncMock()
        conn = AsyncMock()

        desc = [type("D", (), {"name": n})() for n in [
            "invoice_id", "invoice_number", "irn", "direction",
            "transmission_status", "transmission_date",
            "firs_irn", "firs_response_code", "firs_rejection_reason",
            "total_amount", "buyer_name", "issue_date",
        ]]

        cursor = AsyncMock()
        cursor.fetchall = AsyncMock(return_value=[
            ("INV-001", "2026001", "IRN-001", "OUTBOUND",
             "ACCEPTED", date(2026, 3, 1),
             "FIRS-001", "200", None,
             50000.0, "Acme Corp", date(2026, 2, 28)),
            ("INV-002", "2026002", "IRN-002", "OUTBOUND",
             "FAILED_RETRYABLE", date(2026, 3, 2),
             None, "400", "Invalid TIN",
             30000.0, "Beta Ltd", date(2026, 2, 28)),
        ])
        cursor.description = desc

        # Make execute return cursor for SET search_path, then for SELECT
        call_idx = [0]

        async def mock_exec(q, p=None):
            call_idx[0] += 1
            if "SELECT" in str(q):
                return cursor
            set_cursor = AsyncMock()
            return set_cursor

        conn.execute = mock_exec

        cm = AsyncMock()
        cm.__aenter__ = AsyncMock(return_value=conn)
        cm.__aexit__ = AsyncMock(return_value=False)
        pool.connection = MagicMock(return_value=cm)

        content, ct = await transmission.generate(
            pool, {"date_from": "2026-03-01", "date_to": "2026-03-31"}, "comp-1",
        )

        assert ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert isinstance(content, bytes)

        # Verify valid XLSX
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "S/No"
        # Data should be present
        assert ws.cell(row=2, column=2).value is not None


# ── Customer Generator ───────────────────────────────────────────────────


class TestCustomerGenerator:
    @pytest.mark.asyncio
    async def test_returns_valid_xlsx(self):
        pool = AsyncMock()
        conn = AsyncMock()

        desc = [type("D", (), {"name": n})() for n in [
            "customer_id", "company_name", "tin", "customer_type",
            "status", "compliance_score", "total_invoices",
            "total_lifetime_value", "total_lifetime_tax",
            "last_invoice_date", "last_active_date", "created_at",
        ]]

        cursor = AsyncMock()
        cursor.fetchall = AsyncMock(return_value=[
            ("CUS-001", "Acme Corp", "12345678", "B2B",
             "ACTIVE", 85.0, 50,
             5000000.0, 375000.0,
             date(2026, 3, 15), date(2026, 3, 15), date(2025, 1, 1)),
        ])
        cursor.description = desc

        call_idx = [0]
        async def mock_exec(q, p=None):
            call_idx[0] += 1
            if "SELECT" in str(q):
                return cursor
            return AsyncMock()

        conn.execute = mock_exec

        cm = AsyncMock()
        cm.__aenter__ = AsyncMock(return_value=conn)
        cm.__aexit__ = AsyncMock(return_value=False)
        pool.connection = MagicMock(return_value=cm)

        content, ct = await customer.generate(pool, {}, "comp-1")

        assert "spreadsheetml" in ct
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "S/No"


# ── Audit Trail Generator ───────────────────────────────────────────────


class TestAuditTrailGenerator:
    @pytest.mark.asyncio
    async def test_returns_pdf_content_type(self):
        pool = AsyncMock()
        conn = AsyncMock()

        # Audit events
        event_cursor = AsyncMock()
        event_cursor.fetchall = AsyncMock(return_value=[
            ("aud-1", "invoice.created", "invoice", "inv-1",
             "CREATE", "user-1", ["total_amount"], None,
             datetime(2026, 3, 1, tzinfo=timezone.utc)),
        ])
        event_cursor.description = [type("D", (), {"name": n})() for n in [
            "audit_id", "event_type", "entity_type", "entity_id",
            "action", "actor_id", "changed_fields", "metadata", "created_at",
        ]]

        # Action breakdown
        action_cursor = AsyncMock()
        action_cursor.fetchall = AsyncMock(return_value=[("CREATE", 1)])

        call_idx = [0]
        async def mock_exec(q, p=None):
            call_idx[0] += 1
            q_str = str(q)
            if "GROUP BY action" in q_str:
                return action_cursor
            elif "SELECT" in q_str and "audit_log" in q_str:
                return event_cursor
            return AsyncMock()

        conn.execute = mock_exec

        cm = AsyncMock()
        cm.__aenter__ = AsyncMock(return_value=conn)
        cm.__aexit__ = AsyncMock(return_value=False)
        pool.connection = MagicMock(return_value=cm)

        content, ct = await audit_trail.generate(
            pool,
            {"date_from": "2026-03-01", "date_to": "2026-03-31"},
            "comp-1",
        )

        assert ct == "application/pdf"
        html = content.decode("utf-8", errors="replace")
        assert "Audit Trail Report" in html
        assert "CREATE" in html


# ── Monthly Summary Generator ────────────────────────────────────────────


class TestMonthlySummaryGenerator:
    @pytest.mark.asyncio
    async def test_returns_pdf_with_data(self):
        with patch("src.reports.generators.monthly_summary._overview", new_callable=AsyncMock) as mock_ov, \
             patch("src.reports.generators.monthly_summary._invoices", new_callable=AsyncMock) as mock_inv, \
             patch("src.reports.generators.monthly_summary._compliance", new_callable=AsyncMock) as mock_comp:

            mock_ov.return_value = {
                "total_invoices": 100,
                "total_value": 5000000.0,
                "total_tax": 375000.0,
            }
            mock_inv.return_value = {
                "by_status": {"TRANSMITTED": 80, "COMMITTED": 20},
                "by_direction": {"outbound": 90, "inbound": 10},
                "payment_health": {"paid": 60, "unpaid": 40},
            }
            mock_comp.return_value = {
                "firs_acceptance_rate": 95.0,
                "submitted_count": 80,
                "accepted_count": 76,
                "rejected_count": 4,
                "common_rejection_reasons": [],
            }

            content, ct = await monthly_summary.generate(
                _make_mock_pool(), {"date_from": "2026-03-01"}, "comp-1",
            )

        assert ct == "application/pdf"
        html = content.decode("utf-8", errors="replace")
        assert "Monthly Summary Report" in html
        assert "5,000,000.00" in html
        assert "95.0%" in html
