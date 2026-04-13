"""Tests for Excel parser."""

import io

import pytest

from src.ingestion.parsers.excel import ExcelParser


@pytest.fixture
def parser():
    return ExcelParser()


class TestExcelParser:
    @pytest.mark.asyncio
    async def test_basic_parse(self, parser, sample_xlsx_bytes):
        result = await parser.parse(sample_xlsx_bytes, "test.xlsx")
        assert result.file_type == "excel"
        assert result.metadata.parser_type == "excel"
        assert result.metadata.row_count == 2
        assert len(result.raw_data) == 2
        assert result.raw_data[0]["Invoice No"] == "INV-001"
        assert result.raw_data[1]["Amount"] == 2300.5

    @pytest.mark.asyncio
    async def test_sheet_names(self, parser, sample_xlsx_bytes):
        result = await parser.parse(sample_xlsx_bytes, "test.xlsx")
        assert "Invoices" in result.metadata.sheet_names

    @pytest.mark.asyncio
    async def test_empty_workbook(self, parser):
        import openpyxl
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        result = await parser.parse(buf.getvalue(), "empty.xlsx")
        assert result.metadata.row_count == 0
        assert len(result.red_flags) == 1

    @pytest.mark.asyncio
    async def test_multiple_sheets(self, parser):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append([1, 2])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y"])
        ws2.append([3, 4])
        buf = io.BytesIO()
        wb.save(buf)
        result = await parser.parse(buf.getvalue(), "multi.xlsx")
        assert result.metadata.row_count == 2
        assert len(result.metadata.sheet_names) == 2

    @pytest.mark.asyncio
    async def test_duplicate_headers(self, parser):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Name", "Name"])
        ws.append(["a", "b", "c"])
        buf = io.BytesIO()
        wb.save(buf)
        result = await parser.parse(buf.getvalue(), "dups.xlsx")
        row = result.raw_data[0]
        assert "Name" in row
        assert "Name_1" in row
        assert "Name_2" in row

    @pytest.mark.asyncio
    async def test_file_size_in_metadata(self, parser, sample_xlsx_bytes):
        result = await parser.parse(sample_xlsx_bytes, "test.xlsx")
        assert result.metadata.file_size_bytes == len(sample_xlsx_bytes)
