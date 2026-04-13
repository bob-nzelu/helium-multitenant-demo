"""Excel parser using openpyxl."""

from __future__ import annotations

import io
import time

from src.ingestion.models import FileType, ParseMetadata, ParseResult, RedFlag
from src.ingestion.parsers.base import BaseParser


class ExcelParser(BaseParser):
    """Parse .xlsx files using openpyxl in read-only mode."""

    async def parse(self, content: bytes, filename: str) -> ParseResult:
        import openpyxl

        start = time.monotonic()
        red_flags: list[RedFlag] = []

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            sheet_names = wb.sheetnames
            all_rows: list[dict] = []
            sheet_row_counts: dict[str, int] = {}

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    sheet_row_counts[sheet_name] = 0
                    continue

                # Detect header row: first row with >50% non-empty cells
                header_idx = 0
                for i, row in enumerate(rows):
                    non_empty = sum(1 for c in row if c is not None and str(c).strip())
                    if non_empty > len(row) * 0.5:
                        header_idx = i
                        break

                headers = [
                    str(c).strip() if c is not None else f"column_{j}"
                    for j, c in enumerate(rows[header_idx])
                ]

                # Deduplicate header names
                seen: dict[str, int] = {}
                unique_headers: list[str] = []
                for h in headers:
                    if h in seen:
                        seen[h] += 1
                        unique_headers.append(f"{h}_{seen[h]}")
                    else:
                        seen[h] = 0
                        unique_headers.append(h)

                data_rows = rows[header_idx + 1 :]
                count = 0
                for row in data_rows:
                    # Skip entirely empty rows
                    if all(c is None or str(c).strip() == "" for c in row):
                        continue
                    row_dict = {}
                    for j, val in enumerate(row):
                        if j < len(unique_headers):
                            row_dict[unique_headers[j]] = val
                    row_dict["__sheet__"] = sheet_name
                    all_rows.append(row_dict)
                    count += 1

                sheet_row_counts[sheet_name] = count

            if not all_rows:
                red_flags.append(RedFlag(
                    field_name="content",
                    message="Excel file has no data rows",
                    severity="warning",
                ))

            elapsed = (time.monotonic() - start) * 1000
            return ParseResult(
                file_type=FileType.EXCEL.value,
                raw_data=all_rows,
                metadata=ParseMetadata(
                    parser_type="excel",
                    original_filename=filename,
                    file_size_bytes=len(content),
                    row_count=len(all_rows),
                    sheet_names=sheet_names,
                    has_header=True,
                    duration_ms=elapsed,
                ),
                red_flags=red_flags,
            )
        finally:
            wb.close()
