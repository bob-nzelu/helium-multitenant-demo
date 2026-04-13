"""
WS7: WestMetro Excel Formatter — canonical formatting for all report workbooks.

Implements the full Helium Excel Configuration v1.0 spec:
- Color palette, font rules, two-row header system
- Zebra striping, cell alignment, number formats
- Column widths with overflow guard, freeze panes
- White space system, data area borders

Usage:
    wb = create_workbook()
    ws = add_sheet(wb, "Report Title")
    write_headers(ws, columns, schema_paths=None)
    write_data_rows(ws, rows, column_configs)
    apply_formatting(ws)
    content = save_workbook(wb)
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ── 1. Color Palette ─────────────────────────────────────────────────────

HEADER_BG = "003366"
SUBHEADER_BG = "D9E1F2"
SNO_BG = "D3D3D3"
ALT_ROW = "F2F2F2"
WHITE = "FFFFFF"
BOUNDARY = "D9D9D9"
HEADER_FONT_COLOR = "FFFFFF"
SUBHEADER_FONT_COLOR = "666666"
SNO_FONT_COLOR = "333333"
LINK_BLUE = "0563C1"
ERROR_RED = "FF0000"
BORDER_GREY = "808080"
BLACK = "000000"

# Pre-built fills
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type="solid")
SNO_FILL = PatternFill(start_color=SNO_BG, end_color=SNO_BG, fill_type="solid")
ALT_FILL = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
BOUNDARY_FILL = PatternFill(start_color=BOUNDARY, end_color=BOUNDARY, fill_type="solid")

# Pre-built fonts
HEADER_FONT = Font(size=11, bold=True, color=HEADER_FONT_COLOR)
SUBHEADER_FONT = Font(size=9, italic=True, color=SUBHEADER_FONT_COLOR)
SNO_DATA_FONT = Font(size=9, color=SNO_FONT_COLOR)
DATA_FONT = Font(size=11.5)

# Pre-built borders
THICK_BOTTOM = Border(bottom=Side(style="thick", color=BLACK))
THIN_BORDER_GREY_BOTTOM = Border(bottom=Side(style="thin", color=BORDER_GREY))
THIN_BORDER_GREY_RIGHT = Border(right=Side(style="thin", color=BORDER_GREY))
WHITE_BORDER = Border(
    left=Side(style="thin", color=WHITE),
    right=Side(style="thin", color=WHITE),
    top=Side(style="thin", color=WHITE),
    bottom=Side(style="thin", color=WHITE),
)

# Pre-built alignments
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_INDENT = Alignment(horizontal="left", vertical="center", indent=2)
RIGHT_INDENT = Alignment(horizontal="right", vertical="center", indent=2)

# ── Row heights ──────────────────────────────────────────────────────────

HEADER_ROW_HEIGHT = 45
SUBHEADER_ROW_HEIGHT = 28
DATA_ROW_HEIGHT = 25
DEFAULT_COLUMN_WIDTH = 17


# ── 6. Alignment Categories ─────────────────────────────────────────────

class AlignmentCategory:
    """Column alignment category — maps semantic type to openpyxl Alignment."""

    SEQUENCE = Alignment(horizontal="center", vertical="center")
    ID_SHORT = Alignment(horizontal="center", vertical="center")
    ID_LONG = Alignment(horizontal="left", vertical="center", indent=2)
    MONETARY = Alignment(horizontal="right", vertical="center", indent=2)
    MONETARY_DASH = Alignment(horizontal="center", vertical="center")
    COUNT = Alignment(horizontal="center", vertical="center")
    TAG = Alignment(horizontal="center", vertical="center")
    DATE = Alignment(horizontal="left", vertical="center", indent=2)
    DESCRIPTION = Alignment(horizontal="left", vertical="center", indent=2)
    TEXT_DEFAULT = Alignment(horizontal="left", vertical="center", indent=2)


# ── Column Configuration ─────────────────────────────────────────────────

class ColumnConfig:
    """Configuration for a single column."""

    def __init__(
        self,
        header: str,
        key: str,
        width: int = DEFAULT_COLUMN_WIDTH,
        alignment: Alignment = AlignmentCategory.TEXT_DEFAULT,
        number_format: str | None = None,
        is_sno: bool = False,
        is_monetary: bool = False,
    ):
        self.header = header
        self.key = key
        self.width = width
        self.alignment = alignment
        self.number_format = number_format
        self.is_sno = is_sno
        self.is_monetary = is_monetary


# ── Workbook Lifecycle ───────────────────────────────────────────────────

def create_workbook() -> Workbook:
    """Create a new workbook with default sheet removed."""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_sheet(wb: Workbook, title: str) -> Worksheet:
    """Add a named sheet to the workbook."""
    return wb.create_sheet(title)


def save_workbook(wb: Workbook) -> bytes:
    """Save workbook to bytes (for blob upload or HTTP response)."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Header Writing ───────────────────────────────────────────────────────

def write_headers(
    ws: Worksheet,
    columns: list[ColumnConfig],
    schema_paths: list[str] | None = None,
) -> int:
    """
    Write the header row(s) and return the first data row number.

    If schema_paths is provided: two-row header (Row 1 = names, Row 2 = paths).
    Otherwise: single-row header.
    """
    two_row = schema_paths is not None and len(schema_paths) > 0

    # Row 1: column names
    for col_idx, col_cfg in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_cfg.header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    if two_row:
        # Row 2: schema paths
        for col_idx, path in enumerate(schema_paths, start=1):
            cell = ws.cell(row=2, column=col_idx, value=path)
            cell.fill = SUBHEADER_FILL
            cell.font = SUBHEADER_FONT
            cell.alignment = CENTER_WRAP
            cell.border = THICK_BOTTOM
        ws.row_dimensions[2].height = SUBHEADER_ROW_HEIGHT
        # Row 3: spacer (leave blank)
        return 4  # data starts at row 4
    else:
        return 2  # data starts at row 2


def write_data_rows(
    ws: Worksheet,
    rows: list[dict[str, Any]],
    columns: list[ColumnConfig],
    start_row: int = 4,
) -> int:
    """
    Write data rows with formatting.

    Returns the last row number written (or start_row - 1 if no data).
    """
    two_row_header = start_row == 4

    for row_offset, row_data in enumerate(rows):
        row_num = start_row + row_offset
        is_alt = (row_num - start_row) % 2 == 1

        ws.row_dimensions[row_num].height = DATA_ROW_HEIGHT

        for col_idx, col_cfg in enumerate(columns, start=1):
            value = row_data.get(col_cfg.key)

            # Null handling: display as dash
            if value is None:
                value = "-"

            cell = ws.cell(row=row_num, column=col_idx, value=value)

            # S/No column: always SNO_BG, never alternates
            if col_cfg.is_sno:
                cell.fill = SNO_FILL
                cell.font = SNO_DATA_FONT
                cell.alignment = AlignmentCategory.SEQUENCE
            else:
                # Zebra striping
                cell.fill = ALT_FILL if is_alt else WHITE_FILL
                cell.font = DATA_FONT

                # Monetary dash alignment override
                if col_cfg.is_monetary and value == "-":
                    cell.alignment = AlignmentCategory.MONETARY_DASH
                else:
                    cell.alignment = col_cfg.alignment

            # Number format
            if col_cfg.number_format and value != "-":
                cell.number_format = col_cfg.number_format

    last_row = start_row + len(rows) - 1 if rows else start_row - 1
    return last_row


# ── Post-write Formatting ────────────────────────────────────────────────

def apply_column_widths(
    ws: Worksheet,
    columns: list[ColumnConfig],
    data_row_count: int,
    start_row: int = 4,
) -> None:
    """Set column widths with overflow guard (sample first 50 rows)."""
    sample_limit = min(data_row_count, 50)

    for col_idx, col_cfg in enumerate(columns, start=1):
        base_width = col_cfg.width
        header_width = len(col_cfg.header) + 4

        # Overflow guard: scan sampled rows
        content_width = base_width
        for row_offset in range(sample_limit):
            cell = ws.cell(row=start_row + row_offset, column=col_idx)
            if cell.value is not None:
                w = len(str(cell.value)) * 1.1 + 4
                content_width = max(content_width, w)

        # Cap at 80
        final_width = max(base_width, header_width, min(content_width, 80))
        ws.column_dimensions[get_column_letter(col_idx)].width = final_width


def apply_freeze_panes(ws: Worksheet, start_row: int, columns: list[ColumnConfig]) -> None:
    """Freeze panes: columns up to last identifier + header rows."""
    # Determine freeze column: if first col is S/No and second is an ID, freeze at C
    freeze_col = "C" if len(columns) >= 2 else "B"
    ws.freeze_panes = f"{freeze_col}{start_row}"


def apply_data_borders(
    ws: Worksheet,
    last_row: int,
    last_col: int,
) -> None:
    """Apply thin grey border to bottom of last data row and right of last column."""
    if last_row < 1 or last_col < 1:
        return

    # Bottom edge of last data row
    for c in range(1, last_col + 1):
        cell = ws.cell(row=last_row, column=c)
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=existing.right,
            top=existing.top,
            bottom=Side(style="thin", color=BORDER_GREY),
        )

    # Right edge of last data column
    for r in range(1, last_row + 1):
        cell = ws.cell(row=r, column=last_col)
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=Side(style="thin", color=BORDER_GREY),
            top=existing.top,
            bottom=existing.bottom,
        )


def apply_white_space(
    ws: Worksheet,
    last_row: int,
    last_col: int,
) -> None:
    """Paint white space zone + boundary row/column beyond data area."""
    white_rows = 50
    white_cols = 10

    # White zone below data
    for r in range(last_row + 1, last_row + white_rows + 1):
        for c in range(1, last_col + white_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = WHITE_FILL
            cell.border = WHITE_BORDER

    # White zone right of data
    for c in range(last_col + 1, last_col + white_cols + 1):
        for r in range(1, last_row + white_rows + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = WHITE_FILL
            cell.border = WHITE_BORDER

    # Boundary row
    boundary_row = last_row + white_rows + 1
    for c in range(1, last_col + white_cols + 2):
        cell = ws.cell(row=boundary_row, column=c)
        cell.fill = BOUNDARY_FILL
        cell.border = WHITE_BORDER

    # Boundary column
    boundary_col = last_col + white_cols + 1
    for r in range(1, last_row + white_rows + 2):
        cell = ws.cell(row=r, column=boundary_col)
        cell.fill = BOUNDARY_FILL
        cell.border = WHITE_BORDER


def format_sheet_complete(
    ws: Worksheet,
    columns: list[ColumnConfig],
    rows: list[dict[str, Any]],
    schema_paths: list[str] | None = None,
) -> None:
    """
    Complete formatting pipeline for a single sheet.

    1. Write headers
    2. Write data rows
    3. Apply column widths (with overflow guard)
    4. Apply freeze panes
    5. Apply data area borders
    6. Apply white space system
    """
    start_row = write_headers(ws, columns, schema_paths)
    last_row = write_data_rows(ws, rows, columns, start_row)
    last_col = len(columns)

    apply_column_widths(ws, columns, len(rows), start_row)
    apply_freeze_panes(ws, start_row, columns)
    apply_data_borders(ws, last_row, last_col)
    apply_white_space(ws, last_row, last_col)
